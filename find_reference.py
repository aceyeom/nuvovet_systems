import csv
import json
import random
import re
import time
import xml.etree.ElementTree as ET
from pathlib import Path

import requests
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# --- [설정 세션] ---
NCBI_EMAIL = "donghyun040720@gmail.com"
NCBI_TOOL = "nuvovet_reference_sampler"
NCBI_EUTILS_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"

DATA_DIR = Path(__file__).resolve().parent / "backend" / "data" / "converted"
JCR_XLSX_PATH = Path(__file__).resolve().parent / "2025JCRIMPACTFACTORS.xlsx"
SCIMAGO_CSV_PATH = Path(__file__).resolve().parent / "scimagojr 2024.csv"
TEXT_OUTPUT_FILENAME = "test_pmc_references.txt"
JSON_OUTPUT_FILENAME = "test_pmc_references.json"
RANDOM_SAMPLE_SIZE = 10
SEARCH_CANDIDATE_COUNT = 12
SELECT_REFERENCE_COUNT = 2
API_DELAY_SECONDS = 0.5
REQUEST_TIMEOUT_SECONDS = 20
JCR_METRIC_YEAR = "2025"
SCIMAGO_METRIC_YEAR = "2024"

DOG_CAT_SPECIES_TERMS = {
    "dog",
    "dogs",
    "canine",
    "cat",
    "cats",
    "feline",
}

DIRECT_RELEVANCE_TERMS = {
    "pharmacokinetic",
    "pharmacokinetics",
    "pharmacodynamic",
    "pharmacology",
    "bioavailability",
    "absorption",
    "distribution",
    "metabolism",
    "excretion",
    "clearance",
    "half-life",
    "dosage",
    "dose",
    "dosing",
    "efficacy",
    "safety",
    "toxicity",
    "adverse",
    "contraindication",
    "warning",
    "drug interaction",
    "interaction",
    "treatment",
    "label",
    "approved",
    "nada",
    "package insert",
}

FDA_SIGNAL_TERMS = {
    "label",
    "approved",
    "nada",
    "package insert",
}

DDI_SIGNAL_TERMS = {
    "drug interaction",
    "interaction",
}

SOLVENT_EXCLUSION_TERMS = {
    "vehicle control",
    "used as vehicle",
    "vehicle-treated",
    "solvent",
    "solvents",
    "excipient",
    "reagent",
    "stock solution",
    "dissolved in",
    "diluted in dmso",
    "containing 2% dimethyl sulfoxide",
}

METHOD_EXCLUSION_TERMS = {
    "transcriptome",
    "dataset",
    "screening",
    "organoid",
    "cell line",
    "metabolomics",
    "proteomics",
    "rna-seq",
    "high-throughput",
    "acute myeloid leukemia",
    "cell signatures",
}

FDA_LABEL_FIELDS = [
    "drug_identity.brand_names",
    "drug_identity.dosage_form",
    "drug_identity.available_strengths",
    "drug_identity.formulary_status",
    "contraindications",
    "precautions",
    "dosage_and_kinetics.dog.dosage_list",
    "dosage_and_kinetics.cat.dosage_list",
    "section_1_2_10.indications",
    "section_1_2_10.client_info",
    "storage_and_forms.forms",
]


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def normalize_issn(issn: str) -> str:
    return re.sub(r"[^0-9Xx]+", "", str(issn or "").upper())


def split_issns(raw_issn_field: str):
    return [normalized for part in str(raw_issn_field or "").split(",") if (normalized := normalize_issn(part))]


def load_scimago_metrics(csv_path: Path):
    by_issn = {}
    by_title = {}

    if not csv_path.exists():
        return {
            "by_issn": by_issn,
            "by_title": by_title,
            "source_file": str(csv_path),
            "metric_year": SCIMAGO_METRIC_YEAR,
        }

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj, delimiter=";")
        for row in reader:
            title = (row.get("Title") or "").strip()
            normalized_title = normalize_text(title)
            issns = split_issns(row.get("Issn", ""))
            metric_record = {
                "journal_title": title,
                "sourceid": (row.get("Sourceid") or "").strip(),
                "type": (row.get("Type") or "").strip(),
                "issns": issns,
                "sjr": (row.get("SJR") or "").strip(),
                "sjr_best_quartile": (row.get("SJR Best Quartile") or "").strip(),
                "h_index": (row.get("H index") or "").strip(),
                "citations_per_doc_2years": (row.get("Citations / Doc. (2years)") or "").strip(),
                "categories": (row.get("Categories") or "").strip(),
                "areas": (row.get("Areas") or "").strip(),
                "coverage": (row.get("Coverage") or "").strip(),
                "publisher": (row.get("Publisher") or "").strip(),
                "country": (row.get("Country") or "").strip(),
                "metric_year": SCIMAGO_METRIC_YEAR,
                "metric_source": "SCImago Journal Rank",
            }

            if normalized_title and normalized_title not in by_title:
                by_title[normalized_title] = metric_record

            for issn in issns:
                by_issn.setdefault(issn, metric_record)

    return {
        "by_issn": by_issn,
        "by_title": by_title,
        "source_file": str(csv_path),
        "metric_year": SCIMAGO_METRIC_YEAR,
    }


def load_jcr_metrics(xlsx_path: Path):
    by_issn = {}
    by_title = {}

    if not xlsx_path.exists() or load_workbook is None:
        return {
            "by_issn": by_issn,
            "by_title": by_title,
            "source_file": str(xlsx_path),
            "metric_year": JCR_METRIC_YEAR,
        }

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return {
                "by_issn": by_issn,
                "by_title": by_title,
                "source_file": str(xlsx_path),
                "metric_year": JCR_METRIC_YEAR,
            }

        normalized_headers = [str(header).strip() if header is not None else "" for header in headers]
        for row in rows:
            record = dict(zip(normalized_headers, row))
            journal_name = str(record.get("Journal Name") or "").strip()
            abbreviated_journal = str(record.get("Abbreviated Journal") or "").strip()
            issn_values = {
                normalize_issn(record.get("ISSN")),
                normalize_issn(record.get("eISSN")),
            }
            issns = sorted(value for value in issn_values if value)

            metric_record = {
                "journal_title": journal_name,
                "abbreviated_journal": abbreviated_journal,
                "publisher": str(record.get("Publisher") or "").strip(),
                "issns": issns,
                "jif": str(record.get("JIF") or "").strip(),
                "five_year_jif": str(record.get("5-Year JIF") or "").strip(),
                "jif_without_self_cites": str(record.get("JIF Without Self-Cites") or "").strip(),
                "jci": str(record.get("JCI") or "").strip(),
                "jif_quartile": str(record.get("JIF Quartile") or "").strip(),
                "category": str(record.get("Category") or "").strip(),
                "jif_rank": str(record.get("JIF Rank") or "").strip(),
                "total_cites": str(record.get("Total Cites") or "").strip(),
                "total_articles": str(record.get("Total Articles") or "").strip(),
                "citable_items": str(record.get("Citable Items") or "").strip(),
                "metric_year": JCR_METRIC_YEAR,
                "metric_source": "Journal Citation Reports",
            }

            for title_variant in (journal_name, abbreviated_journal):
                normalized_title = normalize_text(title_variant)
                if normalized_title and normalized_title not in by_title:
                    by_title[normalized_title] = metric_record

            for issn in issns:
                by_issn.setdefault(issn, metric_record)
    finally:
        workbook.close()

    return {
        "by_issn": by_issn,
        "by_title": by_title,
        "source_file": str(xlsx_path),
        "metric_year": JCR_METRIC_YEAR,
    }


JCR_METRICS_INDEX = load_jcr_metrics(JCR_XLSX_PATH)
SCIMAGO_METRICS_INDEX = load_scimago_metrics(SCIMAGO_CSV_PATH)


def lookup_jcr_metrics(journal_title: str, journal_issns):
    for issn in journal_issns or []:
        normalized_issn = normalize_issn(issn)
        matched = JCR_METRICS_INDEX["by_issn"].get(normalized_issn)
        if matched:
            return {
                **matched,
                "match_method": "issn",
                "matched_value": normalized_issn,
            }

    normalized_title = normalize_text(journal_title)
    if normalized_title:
        matched = JCR_METRICS_INDEX["by_title"].get(normalized_title)
        if matched:
            return {
                **matched,
                "match_method": "title",
                "matched_value": journal_title,
            }

    return None


def lookup_scimago_metrics(journal_title: str, journal_issns):
    for issn in journal_issns or []:
        normalized_issn = normalize_issn(issn)
        matched = SCIMAGO_METRICS_INDEX["by_issn"].get(normalized_issn)
        if matched:
            return {
                **matched,
                "match_method": "issn",
                "matched_value": normalized_issn,
            }

    normalized_title = normalize_text(journal_title)
    if normalized_title:
        matched = SCIMAGO_METRICS_INDEX["by_title"].get(normalized_title)
        if matched:
            return {
                **matched,
                "match_method": "title",
                "matched_value": journal_title,
            }

    return None


def build_journal_metrics(journal_title: str, journal_issns):
    jcr_metrics = lookup_jcr_metrics(journal_title=journal_title, journal_issns=journal_issns)
    scimago_metrics = lookup_scimago_metrics(journal_title=journal_title, journal_issns=journal_issns)

    primary_metrics = jcr_metrics or scimago_metrics
    if not primary_metrics:
        return None

    metric_summary = {
        "primary_metric_source": primary_metrics.get("metric_source"),
        "match_method": primary_metrics.get("match_method"),
        "matched_value": primary_metrics.get("matched_value"),
        "metric_year": primary_metrics.get("metric_year"),
        "jcr": jcr_metrics,
        "scimago": scimago_metrics,
    }

    if jcr_metrics:
        metric_summary.update(
            {
                "impact_factor": jcr_metrics.get("jif"),
                "impact_factor_5year": jcr_metrics.get("five_year_jif"),
                "impact_factor_quartile": jcr_metrics.get("jif_quartile"),
                "journal_citation_indicator": jcr_metrics.get("jci"),
            }
        )

    if scimago_metrics:
        metric_summary.update(
            {
                "sjr": scimago_metrics.get("sjr"),
                "sjr_best_quartile": scimago_metrics.get("sjr_best_quartile"),
                "h_index": scimago_metrics.get("h_index"),
            }
        )

    return metric_summary


def contains_normalized_term(normalized_text: str, term: str) -> bool:
    normalized_term = normalize_text(term)
    if not normalized_text or not normalized_term:
        return False
    return re.search(rf"(?:^|\s){re.escape(normalized_term)}(?:\s|$)", normalized_text) is not None


def collect_matching_terms(normalized_text: str, terms):
    return sorted(term for term in terms if contains_normalized_term(normalized_text, term))


def collect_species_hits(text: str, terms):
    hits = []
    for term in sorted(terms):
        pattern = re.compile(rf"(?<![A-Za-z]){re.escape(term)}(?![A-Za-z])", re.IGNORECASE)
        for match in pattern.finditer(text):
            matched_text = match.group(0)
            if matched_text.isupper() and len(matched_text) > 1:
                continue
            hits.append(term)
            break
    return hits


def prettify_drug_name(stem: str) -> str:
    name = stem.replace("__", ", ").replace("_", " ").strip()
    name = " ".join(name.split())
    return name.title()


def build_drug_aliases(drug_name: str):
    aliases = {drug_name}
    lowered = drug_name.lower()
    cleaned = re.sub(r"\b(systemic|ophthalmic|topical)\b", "", lowered)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,")
    if cleaned:
        aliases.add(cleaned)
    primary = cleaned.split(",")[0].strip()
    if primary:
        aliases.add(primary)
    return sorted(normalize_text(alias) for alias in aliases if alias.strip())


def iter_drug_records():
    for file_path in sorted(DATA_DIR.rglob("*.jsonl")):
        stem_lower = file_path.stem.lower()
        if "ophthalmic" in stem_lower:
            continue
        yield {
            "name": prettify_drug_name(file_path.stem),
            "file_path": file_path,
        }


def load_drug_payload(file_path: Path):
    with file_path.open("r", encoding="utf-8") as file_obj:
        return json.load(file_obj)


def build_strict_query(drug_name: str) -> str:
    species_block = "(dog[Title/Abstract] OR dogs[Title/Abstract] OR canine[Title/Abstract] OR cat[Title/Abstract] OR cats[Title/Abstract] OR feline[Title/Abstract])"
    relevance_block = (
        '(pharmacokinetic*[Title/Abstract] OR pharmacology[Title/Abstract] OR metabolism[Title/Abstract] '
        'OR clearance[Title/Abstract] OR "half-life"[Title/Abstract] OR bioavailability[Title/Abstract] '
        'OR dosage[Title/Abstract] OR dosing[Title/Abstract] OR safety[Title/Abstract] '
        'OR toxicity[Title/Abstract] OR adverse[Title/Abstract] OR contraindication*[Title/Abstract] '
        'OR "drug interaction"[Title/Abstract] OR efficacy[Title/Abstract] OR label[Title/Abstract] '
        'OR approved[Title/Abstract])'
    )
    exclusion_block = (
        '("vehicle control"[Title/Abstract] OR solvent[Title/Abstract] OR solvents[Title/Abstract] '
        'OR excipient[Title/Abstract] OR reagent[Title/Abstract] OR transcriptome[Title/Abstract] '
        'OR dataset[Title/Abstract] OR "cell line"[Title/Abstract] OR organoid[Title/Abstract] '
        'OR screening[Title/Abstract])'
    )
    return f'"{drug_name}"[Title/Abstract] AND {species_block} AND {relevance_block} NOT {exclusion_block}'


def ncbi_get(endpoint: str, params: dict):
    response = requests.get(
        f"{NCBI_EUTILS_BASE}/{endpoint}",
        params={**params, "email": NCBI_EMAIL, "tool": NCBI_TOOL},
        timeout=REQUEST_TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    return response


def search_pmc_ids(query: str):
    response = ncbi_get(
        "esearch.fcgi",
        {
            "db": "pmc",
            "term": query,
            "retmax": SEARCH_CANDIDATE_COUNT,
            "sort": "relevance",
            "retmode": "json",
        },
    )
    record = response.json()
    return record.get("esearchresult", {}).get("idlist", [])


def fetch_summaries(pmc_ids):
    if not pmc_ids:
        return {}

    response = ncbi_get(
        "esummary.fcgi",
        {
            "db": "pmc",
            "id": ",".join(pmc_ids),
            "retmode": "json",
        },
    )
    result = response.json().get("result", {})

    summary_map = {}
    for pmc_id in result.get("uids", []):
        record = result.get(pmc_id, {})
        pmc_id = str(pmc_id).strip()
        if pmc_id:
            summary_map[pmc_id] = {
                "title": record.get("title", "Title Not Found"),
                "url": f"https://www.ncbi.nlm.nih.gov/pmc/articles/PMC{pmc_id}/",
            }
    return summary_map


def join_text(element):
    if element is None:
        return ""
    return " ".join(part.strip() for part in element.itertext() if part and part.strip())


def fetch_article_metadata(pmc_ids):
    if not pmc_ids:
        return {}

    response = ncbi_get(
        "efetch.fcgi",
        {
            "db": "pmc",
            "id": ",".join(pmc_ids),
            "retmode": "xml",
        },
    )
    raw_xml = response.text

    try:
        root = ET.fromstring(raw_xml)
    except ET.ParseError:
        return {}

    article_map = {}
    for article in root.findall(".//article"):
        pmc_id = join_text(article.find(".//front/article-meta/article-id[@pub-id-type='pmcaid']"))
        if not pmc_id:
            raw_pmcid = join_text(article.find(".//front/article-meta/article-id[@pub-id-type='pmcid']"))
            pmc_id = raw_pmcid.replace("PMC", "").strip()
        abstract_text = join_text(article.find(".//front/article-meta/abstract"))
        journal_title = join_text(article.find(".//front/journal-meta/journal-title-group/journal-title"))
        journal_issns = [
            issn_text
            for issn_text in (
                join_text(node) for node in article.findall(".//front/journal-meta/issn")
            )
            if issn_text
        ]
        if pmc_id:
            article_map[pmc_id] = {
                "abstract": abstract_text,
                "journal_title": journal_title,
                "journal_issns": journal_issns,
            }
    return article_map


def evaluate_candidate(drug_name: str, title: str, abstract_text: str):
    title_abstract = f"{title} {abstract_text}".strip()
    normalized_title_abstract = normalize_text(title_abstract)
    title_lower = title.lower()

    aliases = build_drug_aliases(drug_name)
    drug_hit = any(alias and contains_normalized_term(normalized_title_abstract, alias) for alias in aliases)
    species_hits = collect_species_hits(title_abstract, DOG_CAT_SPECIES_TERMS)
    direct_hits = collect_matching_terms(normalized_title_abstract, DIRECT_RELEVANCE_TERMS)
    solvent_hits = collect_matching_terms(normalized_title_abstract, SOLVENT_EXCLUSION_TERMS)
    method_hits = collect_matching_terms(normalized_title_abstract, METHOD_EXCLUSION_TERMS)

    score = 0
    reasons = []
    excluded = False

    if drug_hit:
        score += 6
        reasons.append("title_or_abstract_contains_drug")
    else:
        excluded = True
        reasons.append("excluded_no_drug_in_title_abstract")

    if species_hits:
        score += 4
        reasons.append(f"species:{', '.join(species_hits[:3])}")
    else:
        excluded = True
        reasons.append("excluded_no_dog_cat_in_title_abstract")

    if direct_hits:
        score += min(len(direct_hits), 5)
        reasons.append(f"relevance:{', '.join(direct_hits[:4])}")

    if any(term in title_lower for term in ["review", "guideline", "consensus", "pharmacokinetic", "toxicity", "safety"]):
        score += 2
        reasons.append("high_signal_title")

    if solvent_hits:
        excluded = True
        reasons.append(f"excluded_solvent:{', '.join(solvent_hits[:3])}")

    if method_hits and len(direct_hits) < 2:
        excluded = True
        reasons.append(f"excluded_method_bias:{', '.join(method_hits[:3])}")

    ddi_relevant = any(contains_normalized_term(normalized_title_abstract, term) for term in DDI_SIGNAL_TERMS)
    fda_relevant = any(contains_normalized_term(normalized_title_abstract, term) for term in FDA_SIGNAL_TERMS)

    return {
        "score": score,
        "reasons": reasons,
        "excluded": excluded,
        "ddi_relevant": ddi_relevant,
        "fda_relevant": fda_relevant,
    }


def get_evidence_targets(file_path: Path, payload: dict):
    data_quality = payload.get("_data_quality") or {}
    pmc_fields = data_quality.get("pmc_rag_fields") or []

    fda_label_fields = list(FDA_LABEL_FIELDS)
    approved_label_signals = []

    dosage_and_kinetics = payload.get("dosage_and_kinetics") or {}
    for species_payload in dosage_and_kinetics.values():
        if not isinstance(species_payload, dict):
            continue
        for dosage in species_payload.get("dosage_list") or []:
            combined = " ".join(
                str(dosage.get(key, ""))
                for key in ("context", "evidence", "duration_note")
            ).lower()
            if "fda" in combined or "label" in combined or "approved" in combined:
                approved_label_signals.append(combined)

    if approved_label_signals:
        fda_label_fields.append("dosage_and_kinetics.*.dosage_list[*].evidence")

    return {
        "pmc_fields": pmc_fields,
        "fda_label_fields": sorted(set(fda_label_fields)),
        "file_path": str(file_path),
    }


def build_metadata_patch(payload: dict, accepted_references, evidence_targets):
    data_quality = payload.get("_data_quality") or {}
    extraction_metadata = payload.get("_extraction_metadata") or {}

    current_ddi_source = data_quality.get("ddi_source")
    current_source_file = extraction_metadata.get("source_file")

    suggested_ddi_source = current_ddi_source
    if accepted_references and any(reference["ddi_relevant"] for reference in accepted_references):
        suggested_ddi_source = "PMC_strict_vet"

    suggested_source_file = current_source_file
    if accepted_references:
        suggested_source_file = f"{current_source_file} + PMC strict vet references"

    return {
        "_data_quality.ddi_source": {
            "current": current_ddi_source,
            "suggested": suggested_ddi_source,
        },
        "_extraction_metadata.source_file": {
            "current": current_source_file,
            "suggested": suggested_source_file,
        },
        "fda_label": {
            "available_now": any(reference["fda_relevant"] for reference in accepted_references),
            "candidate_fields": evidence_targets["fda_label_fields"],
            "status": "pending_manual_label_ingest",
        },
    }


def get_pmc_references(drug_name: str):
    query = build_strict_query(drug_name)

    try:
        pmc_ids = search_pmc_ids(query)
        if not pmc_ids:
            return query, []

        summaries = fetch_summaries(pmc_ids)
        article_metadata = fetch_article_metadata(pmc_ids)

        candidates = []
        for pmc_id in pmc_ids:
            summary = summaries.get(pmc_id, {})
            title = summary.get("title", "Title Not Found")
            article_record = article_metadata.get(pmc_id, {})
            abstract_text = article_record.get("abstract", "")
            journal_title = article_record.get("journal_title", "")
            journal_issns = article_record.get("journal_issns", [])
            evaluation = evaluate_candidate(drug_name=drug_name, title=title, abstract_text=abstract_text)

            if evaluation["excluded"] or evaluation["score"] < 4:
                continue

            journal_metrics = build_journal_metrics(journal_title=journal_title, journal_issns=journal_issns)

            candidates.append(
                {
                    "pmc_id": pmc_id,
                    "title": title,
                    "url": summary.get("url", "URL Not Found"),
                    "abstract": abstract_text,
                    "journal_title": journal_title,
                    "journal_issns": journal_issns,
                    "journal_metrics": journal_metrics,
                    "score": evaluation["score"],
                    "reasons": evaluation["reasons"],
                    "ddi_relevant": evaluation["ddi_relevant"],
                    "fda_relevant": evaluation["fda_relevant"],
                }
            )

        candidates.sort(key=lambda item: item["score"], reverse=True)
        return query, candidates[:SELECT_REFERENCE_COUNT]

    except Exception as error:
        print(f"Error fetching data for {drug_name}: {error}")
        return query, []


def write_text_report(result_rows):
    with open(TEXT_OUTPUT_FILENAME, "w", encoding="utf-8") as file_obj:
        file_obj.write("--- PMC 레퍼런스 수집 테스트 결과 ---\n\n")

        for row in result_rows:
            file_obj.write(f"■ Drug Name: {row['drug_name']}\n")
            file_obj.write(f"  Query: {row['query']}\n")
            file_obj.write(
                "  PMC 보강 필드: "
                + (", ".join(row["evidence_targets"]["pmc_fields"]) if row["evidence_targets"]["pmc_fields"] else "없음")
                + "\n"
            )
            file_obj.write(
                "  FDA 라벨 보강 필드: "
                + ", ".join(row["evidence_targets"]["fda_label_fields"])
                + "\n"
            )

            if not row["accepted_references"]:
                file_obj.write("  -> 제목 또는 초록에 약물명과 개/고양이 종 언급이 동시에 있는 직접 관련 레퍼런스를 찾지 못했습니다.\n")
            else:
                for index, reference in enumerate(row["accepted_references"], start=1):
                    file_obj.write(f"  [{index}] Score: {reference['score']}\n")
                    file_obj.write(f"     Title: {reference['title']}\n")
                    file_obj.write(
                        f"     Journal: {reference.get('journal_title') or 'Unknown Journal'}"
                        + (
                            f" | ISSN: {', '.join(reference.get('journal_issns') or [])}"
                            if reference.get("journal_issns")
                            else ""
                        )
                        + "\n"
                    )
                    journal_metrics = reference.get("journal_metrics") or {}
                    if journal_metrics:
                        if journal_metrics.get("jcr"):
                            jcr_metrics = journal_metrics.get("jcr") or {}
                            file_obj.write(
                                "     JCR: "
                                f"JIF {journal_metrics.get('impact_factor') or 'N/A'}"
                                f", 5Y JIF {journal_metrics.get('impact_factor_5year') or 'N/A'}"
                                f", JCI {journal_metrics.get('journal_citation_indicator') or 'N/A'}"
                                f", Quartile {journal_metrics.get('impact_factor_quartile') or 'N/A'}"
                                f", match={jcr_metrics.get('match_method') or journal_metrics.get('match_method') or 'unknown'}"
                                f", source_year={jcr_metrics.get('metric_year') or journal_metrics.get('metric_year') or 'N/A'}"
                                + "\n"
                            )
                        else:
                            file_obj.write("     JCR: 매핑 없음\n")

                        if journal_metrics.get("scimago"):
                            scimago_metrics = journal_metrics.get("scimago") or {}
                            file_obj.write(
                                "     SCImago: "
                                f"SJR {scimago_metrics.get('sjr') or 'N/A'}"
                                f" ({scimago_metrics.get('sjr_best_quartile') or 'N/A'})"
                                f", H-index {scimago_metrics.get('h_index') or 'N/A'}"
                                f", match={scimago_metrics.get('match_method') or 'unknown'}"
                                f", source_year={scimago_metrics.get('metric_year') or 'N/A'}"
                                + "\n"
                            )
                        else:
                            file_obj.write("     SCImago: 매핑 없음\n")
                    else:
                        file_obj.write("     JCR: 매핑 없음\n")
                        file_obj.write("     SCImago: 매핑 없음\n")
                    file_obj.write(f"     URL: {reference['url']}\n")
                    file_obj.write(f"     Match: {', '.join(reference['reasons'])}\n")

            file_obj.write("-" * 30 + "\n\n")


def write_json_report(result_rows):
    with open(JSON_OUTPUT_FILENAME, "w", encoding="utf-8") as file_obj:
        json.dump(result_rows, file_obj, ensure_ascii=False, indent=2)


def main():
    print("PMC 레퍼런스 수집 테스트를 시작합니다...")

    all_records = list(iter_drug_records())
    sampled_records = random.sample(all_records, min(RANDOM_SAMPLE_SIZE, len(all_records)))
    result_rows = []

    for record in sampled_records:
        drug_name = record["name"]
        print(f"[{drug_name}] 검색 중...")
        time.sleep(API_DELAY_SECONDS)

        payload = load_drug_payload(record["file_path"])
        query, accepted_references = get_pmc_references(drug_name)
        evidence_targets = get_evidence_targets(record["file_path"], payload)
        metadata_patch = build_metadata_patch(payload, accepted_references, evidence_targets)

        result_rows.append(
            {
                "drug_name": drug_name,
                "drug_id": payload.get("id"),
                "drug_file": str(record["file_path"]),
                "query": query,
                "evidence_targets": evidence_targets,
                "accepted_references": accepted_references,
                "proposed_metadata_updates": metadata_patch,
            }
        )

    write_text_report(result_rows)
    write_json_report(result_rows)

    print(
        f"\n테스트가 완료되었습니다. '{TEXT_OUTPUT_FILENAME}' 및 '{JSON_OUTPUT_FILENAME}' 파일을 확인해 주세요."
    )


if __name__ == "__main__":
    main()
